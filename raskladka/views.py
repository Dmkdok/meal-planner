# raskladka/views.py
from flask import (
    Blueprint,
    render_template,
    request,
    jsonify,
    redirect,
    url_for,
    flash,
    send_file,
)
from flask_login import (
    login_user,
    login_required,
    logout_user,
    current_user,
)
from raskladka import db, bcrypt
from raskladka.models import User, MealPlan, Day
from raskladka.services import (
    CalculationService,
    MealPlanService,
    DayService,
    MealService,
    ProductService,
    BackupService,
)
from raskladka.utils import validate_positive_integer
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
from datetime import datetime

views = Blueprint("views", __name__)


def _json_error(message, status_code=400):
    return jsonify({"status": "error", "message": message}), status_code


# Handlers for POST actions on index
def _handle_delete_plan(data):
    try:
        plan_id = int(data.get("plan_id"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректный идентификатор раскладки")
    success = MealPlanService.delete_plan(plan_id, current_user.id)
    if success:
        return jsonify(
            {"status": "success", "redirect": url_for("views.index")}
        )
    return _json_error("Раскладка не найдена или доступ запрещён")


def _handle_update_plan_name(data):
    try:
        plan_id = int(data.get("plan_id"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректный идентификатор раскладки")
    new_name = data.get("new_name")
    success = MealPlanService.update_plan_name(
        plan_id, current_user.id, new_name
    )
    if success:
        return jsonify({"status": "success"})
    return _json_error("Раскладка не найдена или доступ запрещён")


def _handle_delete_day(data):
    try:
        day_id = int(data.get("day_id"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректный идентификатор дня")
    success = DayService.delete_day(day_id, current_user.id)
    if success:
        return jsonify({"status": "success"})
    return _json_error("День не найден или доступ запрещён")


def _handle_update_product(data):
    try:
        product_id = int(data.get("product_id"))
        weight = int(data.get("weight"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректные данные продукта")
    success, message = ProductService.update_product(
        product_id,
        current_user.id,
        data.get("name"),
        weight,
    )
    if success:
        return jsonify({"status": "success"})
    return _json_error(message)


def _handle_delete_product(data):
    try:
        product_id = int(data.get("product_id"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректный идентификатор продукта")
    success = ProductService.delete_product(product_id, current_user.id)
    if success:
        return jsonify({"status": "success"})
    return _json_error("Продукт не найден или доступ запрещён")


def _handle_add_day(data):
    try:
        plan_id = int(data.get("plan_id"))
        day_number = int(data.get("day_number"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректные параметры дня")
    success = DayService.add_day(plan_id, current_user.id, day_number)
    if success:
        return jsonify({"status": "success"})
    return _json_error("Раскладка не найдена или доступ запрещён")


def _handle_add_meal(data):
    try:
        plan_id = int(data.get("plan_id"))
        day_number = int(data.get("day_number"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректные параметры приема пищи")
    success = MealService.add_meal(
        plan_id,
        current_user.id,
        day_number,
        data.get("meal_type"),
    )
    if success:
        return jsonify({"status": "success"})
    return _json_error("День или раскладка не найдены")


def _handle_add_product(data):
    try:
        meal_id = int(data.get("meal_id"))
        weight = int(data.get("weight"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректные данные продукта")
    success, message = ProductService.add_product(
        meal_id,
        current_user.id,
        data.get("name"),
        weight,
    )
    if success:
        return jsonify({"status": "success"})
    return _json_error(message)


def _handle_remove_meal(data):
    try:
        meal_id = int(data.get("meal_id"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректный идентификатор приема пищи")
    success = MealService.delete_meal(meal_id, current_user.id)
    if success:
        return jsonify({"status": "success"})
    return _json_error("Прием пищи не найден или доступ запрещён")


def _handle_update_meal_name(data):
    try:
        meal_id = int(data.get("meal_id"))
    except Exception:  # noqa: BLE001
        return _json_error("Некорректный идентификатор приема пищи")
    success = MealService.update_meal_type(
        meal_id, current_user.id, data.get("meal_name")
    )
    if success:
        return jsonify({"status": "success"})
    return _json_error("Прием пищи не найден или доступ запрещён")


def _handle_create_plan(data):
    new_plan = MealPlan(user_id=current_user.id, name=data.get("name"))
    db.session.add(new_plan)
    db.session.commit()
    return jsonify({"status": "success"})


ACTION_HANDLERS = {
    "delete_plan": _handle_delete_plan,
    "update_plan_name": _handle_update_plan_name,
    "delete_day": _handle_delete_day,
    "update_product": _handle_update_product,
    "delete_product": _handle_delete_product,
    "add_day": _handle_add_day,
    "add_meal": _handle_add_meal,
    "add_product": _handle_add_product,
    "remove_meal": _handle_remove_meal,
    "update_meal_name": _handle_update_meal_name,
    "create_plan": _handle_create_plan,
}


def _validate_export_args():
    """Парсит и валидирует параметры экспорта из query string.

    Returns:
        tuple[str, int, int, Response|None]:
            (plan_id, trip_days, people_count, error_response)
    """
    plan_id = request.args.get("plan_id")
    trip_days = request.args.get("trip_days")
    people_count = request.args.get("people_count")

    if not all([plan_id, trip_days, people_count]):
        return None, None, None, _json_error(
            "Необходимо указать plan_id, trip_days и people_count", 400
        )

    is_valid_trip_days, trip_days_error = validate_positive_integer(
        trip_days, "Количество дней похода"
    )
    if not is_valid_trip_days:
        return None, None, None, _json_error(trip_days_error, 400)

    is_valid_people_count, people_count_error = validate_positive_integer(
        people_count, "Количество человек"
    )
    if not is_valid_people_count:
        return None, None, None, _json_error(people_count_error, 400)

    return plan_id, int(trip_days), int(people_count), None


def _prepare_headers(
    layout_days_count: int,
    meal_types_by_day: list[list[str]],
    people_count: int,
) -> list[str]:
    headers = [
        "Продукт",
        f"1 прием пищи на {people_count} чел.",
    ]
    for i in range(layout_days_count):
        day_meal_types = (
            meal_types_by_day[i] if i < len(meal_types_by_day) else []
        )
        meal_type_text = (
            ", ".join(day_meal_types) if day_meal_types else f"Рацион {i + 1}"
        )
        headers.append(f"Рацион {i + 1}\n{meal_type_text}")
    headers.extend(["Количество повторений", "Общий вес для покупки"])
    return headers


def _apply_table_styles(ws, thin_border):
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            align = "left" if cell.column == 1 else "center"
            cell.alignment = Alignment(horizontal=align, vertical="center")


def _auto_fit_columns(ws):
    column_widths = {}
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            value = str(cell.value) if cell.value is not None else ""
            letter = cell.column_letter
            column_widths[letter] = max(
                column_widths.get(letter, 0),
                len(value) + 2,
            )
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = min(40, width)


def _append_total_row(ws, thin_border, layout_days_count: int):
    if ws.max_row < 2:
        return
    start = ws.cell(row=2, column=ws.max_column).coordinate
    end = ws.cell(row=ws.max_row, column=ws.max_column).coordinate
    sum_cell = f"=SUM({start}:{end})"
    empty_cells = [""] * layout_days_count
    ws.append(["ИТОГО", ""] + empty_cells + ["", sum_cell])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_workbook(
    results: list[dict],
    summary: dict,
    meal_types_by_day: list[list[str]],
    product_meal_usage: dict,
    trip_days: int,
    people_count: int,
):
    layout_days_count = summary["layout_days_count"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Расчет"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="667EEA", end_color="667EEA", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    headers = _prepare_headers(
        layout_days_count, meal_types_by_day, people_count
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    for item in results:
        name = item.get("name", "")
        weight_per_meal = item.get("weight_per_meal", 0) or 0
        weight_for_people = weight_per_meal * people_count

        row = [name, round(weight_for_people, 2)]

        total_occurrences = 0
        usage_map = product_meal_usage.get(name, {})

        for i in range(layout_days_count):
            usage_count = usage_map.get(i, 0)
            if i >= trip_days:
                repetitions = 0
            else:
                repetitions = ((trip_days - 1 - i) // layout_days_count) + 1
            total_usage = usage_count * repetitions
            total_occurrences += total_usage
            row.append(total_usage if total_usage > 0 else "")

        total_weight = round(total_occurrences * weight_for_people, 2)
        row.extend([total_occurrences, total_weight])
        ws.append(row)

    _apply_table_styles(ws, thin_border)
    _auto_fit_columns(ws)
    _append_total_row(ws, thin_border, layout_days_count)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@views.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("views.index"))
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")

        if not username or not password or not confirm_password:
            flash("Все поля обязательны к заполнению", "error")
            return redirect(url_for("views.register"))

        if password != confirm_password:
            flash("Пароль и подтверждение не совпадают", "error")
            return redirect(url_for("views.register"))

        if len(password) < 6:
            flash("Пароль должен быть не короче 6 символов", "error")
            return redirect(url_for("views.register"))

        if User.query.filter_by(username=username).first():
            flash("Имя пользователя уже занято", "error")
            return redirect(url_for("views.register"))

        hashed_password = (
            bcrypt.generate_password_hash(password).decode("utf-8")
        )
        user = User(username=username, password=hashed_password)
        db.session.add(user)
        db.session.commit()
        flash("Регистрация прошла успешно! Теперь вы можете войти.", "success")
        return redirect(url_for("views.login"))

    return render_template("register.html")


@views.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("views.index"))
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        user = User.query.filter_by(username=username).first()
        if user and bcrypt.check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for("views.index"))
        flash("Неверное имя пользователя или пароль", "error")

    return render_template("login.html")


@views.route("/", methods=["GET", "POST"])
@login_required
def index():
    if request.method == "POST":
        data = request.get_json() or {}
        action = data.get("action")
        selected_plan_id = data.get("plan_id")

        try:
            handler = ACTION_HANDLERS.get(action)
            if not handler:
                return _json_error("Неизвестное действие")
            return handler(data)
        except Exception as e:  # noqa: BLE001
            db.session.rollback()
            print(f"Ошибка при обработке запроса: {e}")
            return jsonify({
                "status": "error",
                "message": "Внутренняя ошибка сервера",
            }), 500

    meal_plans = MealPlanService.get_user_plans(current_user.id)
    selected_plan_id = request.args.get("plan_id")
    if selected_plan_id is not None:
        try:
            selected_plan_id = int(selected_plan_id)
        except Exception:  # noqa: BLE001
            selected_plan_id = None
    selected_plan = (
        MealPlanService.get_plan_by_id(selected_plan_id, current_user.id)
        if selected_plan_id
        else meal_plans[0]
        if meal_plans
        else None
    )

    if not selected_plan:
        selected_plan = MealPlanService.create_default_plan(current_user.id)
        meal_plans = [selected_plan]

    return render_template(
        "index.html",
        meal_plans=meal_plans,
        selected_plan=selected_plan,
    )


@views.route("/calculate", methods=["POST"])
@login_required
def calculate_products():
    """API endpoint для расчета продуктов на основе раскладки"""
    try:
        data = request.get_json()
        try:
            plan_id = int(data.get("plan_id"))
        except Exception:  # noqa: BLE001
            return jsonify({
                "status": "error",
                "message": "Некорректный идентификатор раскладки",
            }), 400
        trip_days = data.get("trip_days")
        people_count = data.get("people_count")

        if not all([plan_id, trip_days, people_count]):
            return jsonify(
                {
                    "status": "error",
                    "message": (
                        "Необходимо указать plan_id, trip_days и people_count"
                    ),
                }
            ), 400

        # Валидация входных данных
        is_valid_trip_days, trip_days_error = validate_positive_integer(
            trip_days, "Количество дней похода"
        )
        if not is_valid_trip_days:
            return jsonify(
                {"status": "error", "message": trip_days_error}
            ), 400

        is_valid_people_count, people_count_error = validate_positive_integer(
            people_count, "Количество человек"
        )
        if not is_valid_people_count:
            return jsonify(
                {"status": "error", "message": people_count_error}
            ), 400

        # Получаем план питания
        meal_plan = MealPlanService.get_plan_by_id(plan_id, current_user.id)
        if not meal_plan:
            return jsonify(
                {
                    "status": "error",
                    "message": (
                        "Раскладка не найдена или доступ запрещён"
                    ),
                }
            ), 404

        # Выполняем расчет
        result = CalculationService.calculate_products_from_layout(
            meal_plan, trip_days, people_count
        )

        if not result.get("success"):
            return jsonify(
                {
                    "status": "error",
                    "message": result.get("error", "Ошибка расчета"),
                }
            ), 400

        return jsonify({"status": "success", "data": result})

    except Exception as e:  # noqa: BLE001
        print(f"Ошибка при расчете продуктов: {e}")
        return jsonify(
            {"status": "error", "message": "Внутренняя ошибка сервера"}
        ), 500


@views.route("/day/<int:day_id>/edit")
@login_required
def edit_day(day_id):
    day = db.session.get(Day, day_id)
    if not day or day.meal_plan.user_id != current_user.id:
        flash("День не найден или доступ запрещён")
        return redirect(url_for("views.index"))

    return render_template("edit_day.html", day=day, meal_plan=day.meal_plan)


@views.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("views.login"))


@views.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    if request.method == "POST":
        form_action = request.form.get("action", "change_password")

        if form_action == "change_password":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")

            # Basic validation
            if (
                not current_password
                or not new_password
                or not confirm_password
            ):
                flash("Все поля обязательны к заполнению")
                return redirect(url_for("views.profile"))

            if not bcrypt.check_password_hash(
                current_user.password, current_password
            ):
                flash("Текущий пароль неверный")
                return redirect(url_for("views.profile"))

            if new_password != confirm_password:
                flash("Новый пароль и подтверждение не совпадают")
                return redirect(url_for("views.profile"))

            if len(new_password) < 6:
                flash("Новый пароль должен быть не короче 6 символов")
                return redirect(url_for("views.profile"))

            try:
                hashed_password = (
                    bcrypt.generate_password_hash(new_password).decode("utf-8")
                )
                # Обновляем пароль текущего пользователя
                user = db.session.get(User, current_user.id)
                user.password = hashed_password
                db.session.commit()
                flash("Пароль успешно обновлен")
                return redirect(url_for("views.profile"))
            except Exception as e:  # noqa: BLE001
                db.session.rollback()
                print(f"Ошибка при смене пароля: {e}")
                flash("Не удалось обновить пароль. Попробуйте позже")
                return redirect(url_for("views.profile"))
        else:
            flash("Неизвестное действие")
            return redirect(url_for("views.profile"))

    return render_template("profile.html", user=current_user)


@views.route("/export_excel", methods=["GET"])
@login_required
def export_excel():
    """Экспорт таблицы расчета в Excel"""
    try:
        plan_id = request.args.get("plan_id")
        try:
            plan_id = int(plan_id) if plan_id is not None else None
        except Exception:  # noqa: BLE001
            return _json_error("Некорректный идентификатор раскладки", 400)
        trip_days = request.args.get("trip_days")
        people_count = request.args.get("people_count")

        if not all([plan_id, trip_days, people_count]):
            return _json_error(
                "Необходимо указать plan_id, trip_days и people_count", 400
            )

        # Валидация входных данных
        is_valid_trip_days, trip_days_error = validate_positive_integer(
            trip_days, "Количество дней похода"
        )
        if not is_valid_trip_days:
            return _json_error(trip_days_error, 400)

        is_valid_people_count, people_count_error = validate_positive_integer(
            people_count, "Количество человек"
        )
        if not is_valid_people_count:
            return _json_error(people_count_error, 400)

        trip_days = int(trip_days)
        people_count = int(people_count)

        # Получаем план питания и проверяем доступ
        meal_plan = MealPlanService.get_plan_by_id(plan_id, current_user.id)
        if not meal_plan:
            return _json_error("Раскладка не найдена или доступ запрещён", 404)

        # Рассчитываем данные
        calc_result = CalculationService.calculate_products_from_layout(
            meal_plan, trip_days, people_count
        )
        if not calc_result.get("success"):
            return _json_error(
                calc_result.get("error", "Ошибка расчета"), 400
            )

        results = calc_result["results"]
        summary = calc_result["summary"]
        meal_types_by_day = calc_result["meal_types_by_day"]
        product_meal_usage = calc_result["product_meal_usage"]

        output = _build_workbook(
            results,
            summary,
            meal_types_by_day,
            product_meal_usage,
            trip_days,
            people_count,
        )

        safe_name = meal_plan.name.replace("/", "-").replace("\\", "-")
        filename = f"{safe_name} - расчет.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype=(
                "application/"
                "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    except Exception as e:  # noqa: BLE001
        print(f"Ошибка при экспорте Excel: {e}")
        return _json_error("Внутренняя ошибка сервера", 500)


@views.route("/backup/export", methods=["GET"])
@login_required
def backup_export():
    """Экспорт всех раскладок пользователя в JSON-файл."""
    try:
        data = BackupService.export_user_data(current_user.id)

        # Подготавливаем отдачу файла
        output = BytesIO()
        output.write(
            json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        )
        output.seek(0)

        safe_username = (
            str(current_user.username).replace("/", "-").replace("\\", "-")
        )
        ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        filename = f"raskladka_backup_{safe_username}_{ts}.json"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/json; charset=utf-8",
        )
    except Exception as e:  # noqa: BLE001
        print(f"Ошибка при экспорте JSON: {e}")
        return _json_error("Не удалось создать резервную копию", 500)


@views.route("/backup/import", methods=["POST"])
@login_required
def backup_import():
    """Импорт раскладок пользователя из загруженного JSON-файла."""
    try:
        if "backup_file" not in request.files:
            flash("Файл не загружен")
            return redirect(url_for("views.profile"))

        file = request.files["backup_file"]
        if not file or file.filename == "":
            flash("Выберите файл для загрузки")
            return redirect(url_for("views.profile"))

        # Читаем содержимое файла как UTF-8
        try:
            payload = file.read().decode("utf-8")
            data = json.loads(payload)
        except Exception:  # noqa: BLE001
            flash("Некорректный JSON в файле")
            return redirect(url_for("views.profile"))

        replace = request.form.get("replace", "on") == "on"

        success, message = BackupService.import_user_data(
            current_user.id, data, replace=replace
        )
        if success:
            return redirect(url_for("views.index", import_success="1"))
        else:
            flash(message)
            return redirect(url_for("views.profile"))

    except Exception as e:  # noqa: BLE001
        print(f"Ошибка при импорте JSON: {e}")
        flash("Не удалось импортировать резервную копию")
        return redirect(url_for("views.profile"))
